from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, session, flash
import os
from datetime import datetime, timedelta
import csv
import io
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import PostgreSQL adapter
import psycopg2
from psycopg2.extras import RealDictCursor
from urllib.parse import urlparse

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "rahmonapay-secret-key-2025")

# Parse DATABASE_URL from environment
DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

DEFAULT_BUDGET = 0
NEWSLETTER_FILE = "newsletter_subscribers.xlsx"


def get_db_connection():
    """Get a database connection"""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None


def init_db():
    """Initialize the database with all tables"""
    try:
        conn = get_db_connection()
        if not conn:
            print("✗ Could not connect to database")
            return False
            
        cur = conn.cursor()
        
        # Users table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(255) UNIQUE NOT NULL,
                email VARCHAR(255) UNIQUE NOT NULL,
                phone VARCHAR(50),
                password VARCHAR(255) NOT NULL,
                newsletter_subscribed INTEGER DEFAULT 0,
                sms_alerts INTEGER DEFAULT 0,
                email_alerts INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Expenses table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                budget_type VARCHAR(50) NOT NULL,
                category VARCHAR(100) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                date DATE NOT NULL,
                description TEXT
            )
        """)
        
        # Settings table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                budget_type VARCHAR(50) NOT NULL,
                budget DECIMAL(10, 2) NOT NULL DEFAULT 0,
                UNIQUE(user_id, budget_type)
            )
        """)
        
        # Alerts table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS alerts (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                budget_type VARCHAR(50) NOT NULL,
                alert_type VARCHAR(50) NOT NULL,
                message TEXT NOT NULL,
                sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        conn.commit()
        cur.close()
        conn.close()
        print("✓ Database initialized successfully")
        return True
    except Exception as e:
        print(f"✗ Database error: {e}")
        return False


def init_newsletter_spreadsheet():
    """Initialize the newsletter subscribers spreadsheet"""
    try:
        if not os.path.exists(NEWSLETTER_FILE):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Newsletter Subscribers"
            
            headers = ['ID', 'Username', 'Email', 'Signup Date', 'Status']
            sheet.append(headers)
            
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="38bdf8", end_color="38bdf8", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 30
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 12
            
            wb.save(NEWSLETTER_FILE)
            print(f"✓ Newsletter spreadsheet created: {NEWSLETTER_FILE}")
    except Exception as e:
        print(f"✗ Error creating newsletter spreadsheet: {e}")


def add_newsletter_subscriber(user_id, username, email):
    """Add a new newsletter subscriber to the Excel spreadsheet"""
    try:
        if os.path.exists(NEWSLETTER_FILE):
            wb = load_workbook(NEWSLETTER_FILE)
            sheet = wb.active
        else:
            init_newsletter_spreadsheet()
            wb = load_workbook(NEWSLETTER_FILE)
            sheet = wb.active
        
        signup_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([user_id, username, email, signup_date, "Active"])
        wb.save(NEWSLETTER_FILE)
        print(f"✓ Added newsletter subscriber: {email}")
        return True
    except Exception as e:
        print(f"✗ Error adding newsletter subscriber: {e}")
        return False


def send_email_alert(user_email, username, alert_type, budget_type, total_spent, budget, remaining):
    """Send email alert for budget status"""
    try:
        if alert_type == "50_percent":
            subject = f"⚠️ RahmonaPay Alert: 50% Budget Reached ({budget_type.title()})"
            message = f"""
Hello {username},

You've reached 50% of your {budget_type} budget!

Budget: GH₵ {budget:.2f}
Spent: GH₵ {total_spent:.2f}
Remaining: GH₵ {remaining:.2f}

Consider monitoring your spending to stay within budget.

Best regards,
RahmonaPay Team
"""
        else:
            subject = f"🚨 RahmonaPay Alert: Budget Exceeded ({budget_type.title()})"
            message = f"""
Hello {username},

WARNING: You have exceeded your {budget_type} budget!

Budget: GH₵ {budget:.2f}
Spent: GH₵ {total_spent:.2f}
Over by: GH₵ {abs(remaining):.2f}

Please review your expenses and adjust your spending.

Best regards,
RahmonaPay Team
"""
        
        print(f"📧 EMAIL ALERT to {user_email}:")
        print(f"   Subject: {subject}")
        return True
    except Exception as e:
        print(f"✗ Error sending email alert: {e}")
        return False


def send_sms_alert(phone, username, alert_type, budget_type, total_spent, budget, remaining):
    """Send SMS alert for budget status"""
    try:
        formatted_phone = phone.replace(" ", "").replace("-", "")
        if not formatted_phone.startswith("+"):
            formatted_phone = "+233" + formatted_phone.lstrip("0")
        
        if alert_type == "50_percent":
            message = f"RahmonaPay Alert: You've reached 50% of your {budget_type} budget. Spent: GH₵{total_spent:.2f} / GH₵{budget:.2f}"
        else:
            message = f"RahmonaPay Alert: Budget exceeded! Spent: GH₵{total_spent:.2f} / GH₵{budget:.2f}. Over by: GH₵{abs(remaining):.2f}"
        
        print(f"📱 SMS ALERT to {formatted_phone}:")
        print(f"   Message: {message}")
        return True
    except Exception as e:
        print(f"✗ Error sending SMS alert: {e}")
        return False


def check_and_send_alerts(user_id, budget_type):
    """Check budget status and send alerts if needed"""
    try:
        conn = get_db_connection()
        if not conn:
            return
            
        cur = conn.cursor()
        
        cur.execute("""
            SELECT username, email, phone, email_alerts, sms_alerts 
            FROM users WHERE id = %s
        """, (user_id,))
        user = cur.fetchone()
        
        if not user:
            cur.close()
            conn.close()
            return
        
        username, email, phone, email_alerts, sms_alerts = user
        
        cur.execute("""
            SELECT budget FROM settings 
            WHERE user_id = %s AND budget_type = %s
        """, (user_id, budget_type))
        budget_result = cur.fetchone()
        budget = float(budget_result[0]) if budget_result else 0
        
        if budget <= 0:
            cur.close()
            conn.close()
            return
        
        cur.execute("""
            SELECT COALESCE(SUM(amount), 0) FROM expenses 
            WHERE user_id = %s AND budget_type = %s
        """, (user_id, budget_type))
        total_spent = float(cur.fetchone()[0] or 0)
        
        remaining = budget - total_spent
        percentage_spent = (total_spent / budget) * 100
        
        alert_type = None
        if total_spent > budget:
            alert_type = "over_budget"
        elif percentage_spent >= 50:
            alert_type = "50_percent"
        
        if alert_type:
            cur.execute("""
                SELECT id FROM alerts 
                WHERE user_id = %s 
                AND budget_type = %s 
                AND alert_type = %s 
                AND sent_at > NOW() - INTERVAL '1 day'
            """, (user_id, budget_type, alert_type))
            
            recent_alert = cur.fetchone()
            
            if not recent_alert:
                if email_alerts:
                    send_email_alert(email, username, alert_type, budget_type, 
                                   total_spent, budget, remaining)
                
                if sms_alerts and phone:
                    send_sms_alert(phone, username, alert_type, budget_type, 
                                 total_spent, budget, remaining)
                
                message = f"Alert sent: {alert_type} for {budget_type} budget"
                cur.execute("""
                    INSERT INTO alerts (user_id, budget_type, alert_type, message) 
                    VALUES (%s, %s, %s, %s)
                """, (user_id, budget_type, alert_type, message))
                
                conn.commit()
                print(f"✓ Alert sent to user {username} for {budget_type} budget")
        
        cur.close()
        conn.close()
        
    except Exception as e:
        print(f"✗ Error checking/sending alerts: {e}")


def get_expenses():
    """Retrieve all expenses from database"""
    if 'user_id' not in session or 'budget_type' not in session:
        return []
    
    try:
        conn = get_db_connection()
        if not conn:
            return []
            
        cur = conn.cursor()
        cur.execute("""
            SELECT category, amount, date, id, description 
            FROM expenses 
            WHERE user_id = %s AND budget_type = %s 
            ORDER BY date DESC, id DESC
        """, (session['user_id'], session['budget_type']))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return rows
    except Exception as e:
        print(f"Error fetching expenses: {e}")
        return []


def get_budget():
    """Get the current budget from settings"""
    if 'user_id' not in session or 'budget_type' not in session:
        return DEFAULT_BUDGET
    
    try:
        conn = get_db_connection()
        if not conn:
            return DEFAULT_BUDGET
            
        cur = conn.cursor()
        cur.execute("""
            SELECT budget FROM settings 
            WHERE user_id = %s AND budget_type = %s
        """, (session['user_id'], session['budget_type']))
        result = cur.fetchone()
        cur.close()
        conn.close()
        return float(result[0]) if result else DEFAULT_BUDGET
    except Exception as e:
        print(f"Error fetching budget: {e}")
        return DEFAULT_BUDGET


def set_budget(amount):
    """Update the budget in settings"""
    if 'user_id' not in session or 'budget_type' not in session:
        return False
    
    try:
        conn = get_db_connection()
        if not conn:
            return False
            
        cur = conn.cursor()
        
        cur.execute("""
            SELECT id FROM settings 
            WHERE user_id = %s AND budget_type = %s
        """, (session['user_id'], session['budget_type']))
        exists = cur.fetchone()
        
        if exists:
            cur.execute("""
                UPDATE settings SET budget = %s 
                WHERE user_id = %s AND budget_type = %s
            """, (amount, session['user_id'], session['budget_type']))
        else:
            cur.execute("""
                INSERT INTO settings (user_id, budget_type, budget) 
                VALUES (%s, %s, %s)
            """, (session['user_id'], session['budget_type'], amount))
        
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f"Error updating budget: {e}")
        return False


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function


@app.route("/")
@login_required
def index():
    """Main route - show dashboard landing"""
    try:
        conn = get_db_connection()
        if not conn:
            return render_template("dashboard.html", 
                                 username=session.get('username', 'User'), 
                                 email="", newsletter_subscribed=0,
                                 phone="", email_alerts=1, sms_alerts=0)
        
        cur = conn.cursor()
        cur.execute("""
            SELECT username, email, newsletter_subscribed, phone, email_alerts, sms_alerts 
            FROM users WHERE id = %s
        """, (session['user_id'],))
        user = cur.fetchone()
        cur.close()
        conn.close()
        
        if user:
            return render_template("dashboard.html", 
                                 username=user[0], email=user[1],
                                 newsletter_subscribed=user[2],
                                 phone=user[3] or "", email_alerts=user[4],
                                 sms_alerts=user[5])
        else:
            return render_template("dashboard.html", 
                                 username=session.get('username', 'User'), 
                                 email="", newsletter_subscribed=0,
                                 phone="", email_alerts=1, sms_alerts=0)
    except Exception as e:
        print(f"Error loading dashboard: {e}")
        return render_template("dashboard.html", 
                             username=session.get('username', 'User'), 
                             email="", newsletter_subscribed=0,
                             phone="", email_alerts=1, sms_alerts=0)


@app.route("/tracker/<budget_type>")
@login_required
def tracker(budget_type):
    """Expense tracker for specific budget type"""
    if budget_type not in ['daily', 'weekly', 'monthly', 'annually']:
        return redirect(url_for('index'))
    
    session['budget_type'] = budget_type
    
    budget_names = {
        'daily': 'Daily',
        'weekly': 'Weekly',
        'monthly': 'Monthly',
        'annually': 'Annual'
    }
    
    return render_template("form.html", 
                         budget_type=budget_type,
                         budget_name=budget_names[budget_type])


@app.route("/update-profile", methods=["POST"])
@login_required
def update_profile():
    """Update user profile"""
    try:
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_new_password = request.form.get("confirm_new_password", "")
        newsletter_subscription = request.form.get("newsletter_subscription") == "on"
        email_alerts = request.form.get("email_alerts") == "on"
        sms_alerts = request.form.get("sms_alerts") == "on"
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "error")
            return redirect(url_for('index'))
            
        cur = conn.cursor()
        
        cur.execute("""
            SELECT username, email, newsletter_subscribed 
            FROM users WHERE id = %s
        """, (session['user_id'],))
        user = cur.fetchone()
        current_newsletter_status = user[2] if user else 0
        
        if email:
            cur.execute("UPDATE users SET email = %s WHERE id = %s", 
                       (email, session['user_id']))
        
        if phone:
            cur.execute("UPDATE users SET phone = %s WHERE id = %s", 
                       (phone, session['user_id']))
        
        cur.execute("""
            UPDATE users SET newsletter_subscribed = %s WHERE id = %s
        """, (1 if newsletter_subscription else 0, session['user_id']))
        
        cur.execute("""
            UPDATE users SET email_alerts = %s, sms_alerts = %s WHERE id = %s
        """, (1 if email_alerts else 0, 1 if sms_alerts else 0, session['user_id']))
        
        if newsletter_subscription and not current_newsletter_status:
            username = user[0] if user else session.get('username', 'User')
            user_email = email if email else (user[1] if user else '')
            add_newsletter_subscriber(session['user_id'], username, user_email)
        
        if current_password and new_password:
            cur.execute("SELECT password FROM users WHERE id = %s", (session['user_id'],))
            user = cur.fetchone()
            
            if user and check_password_hash(user[0], current_password):
                if new_password == confirm_new_password:
                    if len(new_password) >= 6:
                        hashed_password = generate_password_hash(new_password)
                        cur.execute("UPDATE users SET password = %s WHERE id = %s", 
                                   (hashed_password, session['user_id']))
                        flash("Password updated successfully!", "success")
                    else:
                        flash("New password must be at least 6 characters", "error")
                        cur.close()
                        conn.close()
                        return redirect(url_for('index'))
                else:
                    flash("New passwords do not match", "error")
                    cur.close()
                    conn.close()
                    return redirect(url_for('index'))
            else:
                flash("Current password is incorrect", "error")
                cur.close()
                conn.close()
                return redirect(url_for('index'))
        
        conn.commit()
        cur.close()
        conn.close()
        
        flash("Profile updated successfully!", "success")
        return redirect(url_for('index'))
        
    except Exception as e:
        print(f"Error updating profile: {e}")
        flash("An error occurred while updating profile", "error")
        return redirect(url_for('index'))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    """Sign up route"""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        newsletter = request.form.get("newsletter") == "on"
        sms_alerts_opt = request.form.get("sms_alerts") == "on"
        
        print(f"Signup attempt - Username: {username}, Email: {email}")
        
        if not username or not email or not password:
            flash("Username, email and password are required", "error")
            return redirect(url_for("signup"))
        
        if password != confirm_password:
            flash("Passwords do not match", "error")
            return redirect(url_for("signup"))
        
        if len(password) < 6:
            flash("Password must be at least 6 characters", "error")
            return redirect(url_for("signup"))
        
        if sms_alerts_opt and not phone:
            flash("Phone number is required for SMS alerts", "error")
            return redirect(url_for("signup"))
        
        try:
            conn = get_db_connection()
            if not conn:
                flash("Database connection error", "error")
                return redirect(url_for("signup"))
                
            cur = conn.cursor()
            
            cur.execute("""
                SELECT id FROM users WHERE username = %s OR email = %s
            """, (username, email))
            if cur.fetchone():
                flash("Username or email already exists", "error")
                cur.close()
                conn.close()
                return redirect(url_for("signup"))
            
            hashed_password = generate_password_hash(password)
            cur.execute("""
                INSERT INTO users (username, email, phone, password, newsletter_subscribed, sms_alerts, email_alerts) 
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (username, email, phone if phone else None, hashed_password, 
                  1 if newsletter else 0, 1 if sms_alerts_opt else 0, 1))
            user_id = cur.fetchone()[0]
            
            for budget_type in ['daily', 'weekly', 'monthly', 'annually']:
                cur.execute("""
                    INSERT INTO settings (user_id, budget_type, budget) 
                    VALUES (%s, %s, %s)
                """, (user_id, budget_type, DEFAULT_BUDGET))
            
            conn.commit()
            cur.close()
            conn.close()
            
            if newsletter:
                add_newsletter_subscriber(user_id, username, email)
            
            print(f"✓ User created successfully: {username} (ID: {user_id})")
            flash("Account created successfully! Please log in.", "success")
            return redirect(url_for("login_page"))
            
        except Exception as e:
            print(f"✗ Error during signup: {e}")
            flash(f"Error: {str(e)}", "error")
            return redirect(url_for("signup"))
    
    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login_page():
    """Login route"""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        
        if not username or not password:
            flash("Username and password are required", "error")
            return redirect(url_for("login_page"))
        
        try:
            conn = get_db_connection()
            if not conn:
                flash("Database connection error", "error")
                return redirect(url_for("login_page"))
                
            cur = conn.cursor()
            cur.execute("""
                SELECT id, password, username 
                FROM users 
                WHERE username = %s OR email = %s
            """, (username, username))
            user = cur.fetchone()
            cur.close()
            conn.close()
            
            if user and check_password_hash(user[1], password):
                session['user_id'] = user[0]
                session['username'] = user[2]
                flash(f"Welcome back, {user[2]}!", "success")
                return redirect(url_for("index"))
            else:
                flash("Invalid username or password", "error")
                return redirect(url_for("login_page"))
                
        except Exception as e:
            print(f"Database error: {e}")
            flash("An error occurred. Please try again.", "error")
            return redirect(url_for("login_page"))
    
    return render_template("login.html")


@app.route("/logout")
def logout():
    """Logout route"""
    username = session.get('username', 'User')
    session.clear()
    flash(f"Goodbye, {username}!", "success")
    return redirect(url_for("login_page"))


@app.route("/add_expense", methods=["POST"])
@login_required
def add_expense():
    """Add expense route"""
    if 'budget_type' not in session:
        return redirect(url_for('index'))
    
    try:
        category = request.form.get("category", "").strip()
        amount_str = request.form.get("amount", "")
        date = request.form.get("date", "")
        description = request.form.get("description", "").strip()

        if not category or not amount_str or not date:
            return redirect(url_for('tracker', budget_type=session['budget_type']))

        try:
            amount = float(amount_str)
            if amount <= 0:
                return redirect(url_for('tracker', budget_type=session['budget_type']))
        except ValueError:
            return redirect(url_for('tracker', budget_type=session['budget_type']))

        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            return redirect(url_for('tracker', budget_type=session['budget_type']))

        conn = get_db_connection()
        if not conn:
            return redirect(url_for('tracker', budget_type=session['budget_type']))
            
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO expenses (user_id, budget_type, category, amount, date, description) 
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (session['user_id'], session['budget_type'], category, amount, date, 
              description if description else None))
        conn.commit()
        cur.close()
        conn.close()
        
        print(f"✓ Added to {session['budget_type']}: {category} - GH₵{amount}")
        check_and_send_alerts(session['user_id'], session['budget_type'])
        
        return redirect(url_for('tracker', budget_type=session['budget_type']))
        
    except Exception as e:
        print(f"Error: {e}")
        return redirect(url_for('tracker', budget_type=session['budget_type']))


@app.route("/api/expenses", methods=["GET"])
@login_required
def api_expenses():
    """API endpoint to get all expenses and calculations"""
    expenses = get_expenses()
    budget = get_budget()
    
    total_spent = sum(float(e[1]) for e in expenses)
    remaining = budget - total_spent
    over_budget = total_spent > budget
    half_budget = budget / 2
    at_half_limit = total_spent >= half_budget and not over_budget
    
    category_totals = {}
    for cat, amt, _, _, _ in expenses:
        category_totals[cat] = category_totals.get(cat, 0) + float(amt)
    
    # Convert dates to strings for JSON serialization
    expenses_list = []
    for exp in expenses:
        expenses_list.append([
            exp[0],  # category
            float(exp[1]),  # amount
            str(exp[2]),  # date
            exp[3],  # id
            exp[4]  # description
        ])
    
    return jsonify({
        "expenses": expenses_list,
        "total_spent": round(total_spent, 2),
        "remaining": round(abs(remaining), 2),
        "over_budget": over_budget,
        "at_half_limit": at_half_limit,
        "category_totals": {k: round(v, 2) for k, v in category_totals.items()},
        "budget": budget
    })


@app.route("/api/budget", methods=["POST"])
@login_required
def update_budget():
    """API endpoint to update the budget"""
    try:
        data = request.get_json()
        new_budget = float(data.get("budget", 0))
        
        if new_budget <= 0:
            return jsonify({"success": False, "message": "Budget must be greater than 0"}), 400
        
        if set_budget(new_budget):
            print(f"✓ Budget updated to GH₵{new_budget}")
            return jsonify({"success": True, "budget": new_budget})
        else:
            return jsonify({"success": False, "message": "Failed to update budget"}), 500
            
    except Exception as e:
        print(f"Error updating budget: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/delete/<int:expense_id>", methods=["POST"])
@login_required
def delete_expense(expense_id):
    """Delete a specific expense by ID"""
    if 'budget_type' not in session:
        return redirect(url_for('index'))
    
    try:
        conn = get_db_connection()
        if not conn:
            return redirect(url_for('tracker', budget_type=session['budget_type']))
            
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM expenses 
            WHERE id = %s AND user_id = %s AND budget_type = %s
        """, (expense_id, session['user_id'], session['budget_type']))
        conn.commit()
        cur.close()
        conn.close()
        print(f"✓ Deleted expense ID {expense_id}")
    except Exception as e:
        print(f"Error deleting expense: {e}")
    
    return redirect(url_for('tracker', budget_type=session['budget_type']))


@app.route("/edit/<int:expense_id>", methods=["POST"])
@login_required
def edit_expense(expense_id):
    """Edit a specific expense by ID"""
    if 'budget_type' not in session:
        return redirect(url_for('index'))
    
    try:
        category = request.form.get("category", "").strip()
        amount_str = request.form.get("amount", "")
        date = request.form.get("date", "")
        description = request.form.get("description", "").strip()

        if not category or not amount_str or not date:
            return redirect(url_for('tracker', budget_type=session['budget_type']))

        try:
            amount = float(amount_str)
            if amount <= 0:
                return redirect(url_for('tracker', budget_type=session['budget_type']))
        except ValueError:
            return redirect(url_for('tracker', budget_type=session['budget_type']))

        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            return redirect(url_for('tracker', budget_type=session['budget_type']))

        conn = get_db_connection()
        if not conn:
            return redirect(url_for('tracker', budget_type=session['budget_type']))
            
        cur = conn.cursor()
        cur.execute("""
            UPDATE expenses 
            SET category = %s, amount = %s, date = %s, description = %s 
            WHERE id = %s AND user_id = %s AND budget_type = %s
        """, (category, amount, date, description if description else None, 
              expense_id, session['user_id'], session['budget_type']))
        conn.commit()
        cur.close()
        conn.close()
        
        print(f"✓ Updated expense ID {expense_id}")
        
    except Exception as e:
        print(f"Error updating expense: {e}")
    
    return redirect(url_for('tracker', budget_type=session['budget_type']))


@app.route("/export/csv")
@login_required
def export_csv():
    """Export all expenses to CSV file"""
    try:
        expenses = get_expenses()
        budget = get_budget()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['Category', 'Amount (GH₵)', 'Date', 'Description', 'ID'])
        
        for exp in expenses:
            writer.writerow([exp[0], float(exp[1]), str(exp[2]), exp[4] if exp[4] else '', exp[3]])
        
        writer.writerow([])
        writer.writerow(['Summary'])
        writer.writerow(['Total Spent', sum(float(e[1]) for e in expenses)])
        writer.writerow(['Budget', budget])
        writer.writerow(['Remaining', budget - sum(float(e[1]) for e in expenses)])
        
        output.seek(0)
        filename = f"rahmonapay_expenses_{datetime.now().strftime('%Y%m%d')}.csv"
        
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        print(f"✓ Exported {len(expenses)} expenses to CSV")
        
        return send_file(mem, mimetype='text/csv', as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Error exporting CSV: {e}")
        return redirect(url_for("index"))


@app.route("/export/excel")
@login_required
def export_excel():
    """Export all expenses to Excel-compatible CSV"""
    try:
        expenses = get_expenses()
        budget = get_budget()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['RahmonaPay Expense Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])
        
        writer.writerow(['Category', 'Amount (GH₵)', 'Date', 'Description', 'Expense ID'])
        
        for exp in expenses:
            writer.writerow([exp[0], f'{float(exp[1]):.2f}', str(exp[2]), exp[4] if exp[4] else '', exp[3]])
        
        writer.writerow([])
        writer.writerow(['SUMMARY'])
        writer.writerow(['Monthly Budget', f'{budget:.2f}'])
        writer.writerow(['Total Spent', f'{sum(float(e[1]) for e in expenses):.2f}'])
        writer.writerow(['Remaining', f'{budget - sum(float(e[1]) for e in expenses):.2f}'])
        
        writer.writerow([])
        writer.writerow(['CATEGORY BREAKDOWN'])
        category_totals = {}
        for cat, amt, _, _, _ in expenses:
            category_totals[cat] = category_totals.get(cat, 0) + float(amt)
        
        for cat, amt in category_totals.items():
            writer.writerow([cat, f'{amt:.2f}'])
        
        output.seek(0)
        filename = f"rahmonapay_report_{datetime.now().strftime('%Y%m%d')}.csv"
        
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        print(f"✓ Exported {len(expenses)} expenses to Excel format")
        
        return send_file(mem, mimetype='text/csv', as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Error exporting Excel: {e}")
        return redirect(url_for("index"))


@app.route("/download-newsletter-list")
@login_required
def download_newsletter_list():
    """Download the newsletter subscribers Excel file"""
    try:
        if not os.path.exists(NEWSLETTER_FILE):
            flash("Newsletter file not found", "error")
            return redirect(url_for("index"))
        
        return send_file(
            NEWSLETTER_FILE,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"newsletter_subscribers_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    except Exception as e:
        print(f"Error downloading newsletter list: {e}")
        flash("Error downloading newsletter list", "error")
        return redirect(url_for("index"))


@app.route("/api/dashboard-stats")
@login_required
def dashboard_stats():
    """API endpoint to get dashboard statistics"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'total_expenses': 0, 'week_expenses': 0, 'month_expenses': 0,
                'budget_status': 'good', 'category_totals': {}, 'recent_expenses': []
            })
        
        cur = conn.cursor()
        
        cur.execute("""
            SELECT category, amount, date, budget_type 
            FROM expenses 
            WHERE user_id = %s 
            ORDER BY date DESC
        """, (session['user_id'],))
        all_expenses = cur.fetchall()
        
        cur.execute("""
            SELECT budget_type, budget 
            FROM settings 
            WHERE user_id = %s
        """, (session['user_id'],))
        budgets = dict(cur.fetchall())
        
        cur.close()
        conn.close()
        
        total_expenses = sum(float(exp[1]) for exp in all_expenses)
        
        today = datetime.now()
        week_ago = today - timedelta(days=7)
        week_expenses = sum(
            float(exp[1]) for exp in all_expenses 
            if exp[2] >= week_ago.date()
        )
        
        current_month = today.strftime("%Y-%m")
        month_expenses = sum(
            float(exp[1]) for exp in all_expenses 
            if str(exp[2]).startswith(current_month)
        )
        
        total_budget = sum(float(b) for b in budgets.values()) if budgets else 0
        budget_status = 'good'
        
        if total_budget > 0:
            if total_expenses > total_budget:
                budget_status = 'over'
            elif total_expenses >= total_budget * 0.8:
                budget_status = 'warning'
        
        category_totals = {}
        for cat, amt, _, _ in all_expenses:
            category_totals[cat] = category_totals.get(cat, 0) + float(amt)
        
        recent_expenses = [
            {
                'category': exp[0],
                'amount': float(exp[1]),
                'date': str(exp[2]),
                'budget_type': exp[3]
            }
            for exp in all_expenses[:10]
        ]
        
        return jsonify({
            'total_expenses': round(total_expenses, 2),
            'week_expenses': round(week_expenses, 2),
            'month_expenses': round(month_expenses, 2),
            'budget_status': budget_status,
            'category_totals': {k: round(v, 2) for k, v in category_totals.items()},
            'recent_expenses': recent_expenses
        })
        
    except Exception as e:
        print(f"Error getting dashboard stats: {e}")
        return jsonify({
            'total_expenses': 0, 'week_expenses': 0, 'month_expenses': 0,
            'budget_status': 'good', 'category_totals': {}, 'recent_expenses': []
        })


@app.route("/init-database")
def initialize_database():
    """Manual database initialization endpoint"""
    success = init_db()
    init_newsletter_spreadsheet()
    
    if success:
        return """
        <html>
        <head>
            <title>Database Initialized</title>
            <style>
                body {
                    font-family: system-ui, sans-serif;
                    background: #0f172a;
                    color: #e5e7eb;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    height: 100vh;
                    margin: 0;
                }
                .message {
                    text-align: center;
                    padding: 40px;
                    background: #1e293b;
                    border-radius: 20px;
                    border: 2px solid #38bdf8;
                }
                h1 { color: #38bdf8; margin-bottom: 20px; }
                a {
                    display: inline-block;
                    margin-top: 20px;
                    padding: 12px 24px;
                    background: linear-gradient(135deg, #38bdf8, #2563eb);
                    color: white;
                    text-decoration: none;
                    border-radius: 10px;
                    font-weight: 600;
                }
                a:hover { transform: translateY(-2px); }
            </style>
        </head>
        <body>
            <div class="message">
                <h1>✓ Database Initialized Successfully!</h1>
                <p>Your RahmonaPay PostgreSQL database has been set up and is ready to use.</p>
                <a href="/login">Go to Login Page</a>
            </div>
        </body>
        </html>
        """
    else:
        return """
        <html>
        <head>
            <title>Database Error</title>
            <style>
                body {
                    font-family: system-ui, sans-serif;
                    background: #0f172a;
                    color: #e5e7eb;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    height: 100vh;
                    margin: 0;
                }
                .message {
                    text-align: center;
                    padding: 40px;
                    background: #1e293b;
                    border-radius: 20px;
                    border: 2px solid #ef4444;
                }
                h1 { color: #ef4444; margin-bottom: 20px; }
            </style>
        </head>
        <body>
            <div class="message">
                <h1>✗ Error Initializing Database</h1>
                <p>Check your DATABASE_URL environment variable and database connection.</p>
            </div>
        </body>
        </html>
        """


if __name__ == "__main__":
    # Initialize database on startup
    init_db()
    init_newsletter_spreadsheet()
    
    print("=" * 60)
    print("🚀 RahmonaPay Expense Tracker Starting (PostgreSQL)...")
    print("=" * 60)
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host='0.0.0.0', port=port)